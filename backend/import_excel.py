#!/usr/bin/env python3
"""
엑셀 → DB 업무 가져오기 스크립트

엑셀 컬럼 형식: 일 | 요일 | 프로젝트 | TO-DO | ...
시트 이름 형식: 01월, 02월, ... 06월 등 (월 자동 감지)

사용법:
  python import_excel.py --file 파일.xlsx --year 2026            # 전체 시트
  python import_excel.py --file 파일.xlsx --year 2026 --month 6  # 특정 월만
  python import_excel.py --file 파일.xlsx --year 2026 --dry-run  # 미리보기
"""

import sys
import re
import argparse
from datetime import date

import openpyxl

sys.path.insert(0, ".")

from app.database import SessionLocal
from app.models.category import Category
from app.models.task import Task


def find_col(headers: list, candidates: list):
    for candidate in candidates:
        for i, h in enumerate(headers):
            if h and candidate.lower() in str(h).strip().lower():
                return i
    return None


def parse_day(val, year: int, month: int):
    if val is None:
        return None
    if hasattr(val, "year"):
        return val.date() if hasattr(val, "date") else val
    try:
        day = int(float(str(val)))
        if 1 <= day <= 31:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    except (ValueError, TypeError):
        pass
    return None


def detect_month_from_sheet(sheet_name: str):
    m = re.search(r"(\d{1,2})월", sheet_name)
    if m:
        return int(m.group(1))
    return None


def process_sheet(ws, year: int, month: int, cat_map: dict, unmatched_cat):
    # 헤더 찾기
    headers = None
    header_row = 1
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if any(cell is not None for cell in row):
            headers = list(row)
            header_row = row_idx
            break

    if not headers:
        return [], {}

    col_day     = find_col(headers, ["일"])
    col_project = find_col(headers, ["프로젝트", "project", "구분"])
    col_todo    = find_col(headers, ["to-do", "todo", "할일", "할 일", "업무"])

    if col_todo is None:
        return [], {}

    items = []
    unmatched = {}
    last_date = None

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        todo_val = row[col_todo] if col_todo < len(row) else None
        if not todo_val or str(todo_val).strip() == "":
            continue

        title = str(todo_val).strip()

        # 날짜
        day_val = row[col_day] if col_day is not None and col_day < len(row) else None
        task_date = parse_day(day_val, year, month)
        if task_date:
            last_date = task_date
        elif last_date:
            task_date = last_date
        else:
            task_date = date(year, month, 1)

        # 프로젝트 → 업무구분 매칭
        project = ""
        if col_project is not None and col_project < len(row) and row[col_project]:
            project = str(row[col_project]).strip()

        category = cat_map.get(project.lower()) if project else None
        if not category and project:
            unmatched[project] = unmatched.get(project, 0) + 1
            category = unmatched_cat

        items.append({
            "date": task_date,
            "category_obj": category,
            "category_name": category.name if category else "미분류",
            "title": title,
            "project": project,
        })

    return items, unmatched


def main():
    parser = argparse.ArgumentParser(description="엑셀 업무 가져오기")
    parser.add_argument("--file", required=True)
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--month", type=int, default=None, help="특정 월만 (없으면 전체 시트)")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file, data_only=True)
    except FileNotFoundError:
        print(f"❌ 파일을 찾을 수 없습니다: {args.file}")
        sys.exit(1)

    # DB 연결
    db = SessionLocal()
    categories = db.query(Category).order_by(Category.sort_order).all()
    cat_map = {c.name.strip().lower(): c for c in categories}

    print("📂 등록된 업무구분:", [c.name for c in categories])

    # 매칭 안되는 항목은 기타로 저장
    unmatched_cat = cat_map.get("기타") or cat_map.get("미분류")
    if not unmatched_cat:
        print("❌ '기타' 업무구분이 없습니다. 업무구분 관리에서 추가해주세요.")
        db.close()
        sys.exit(1)

    # 처리할 시트 결정
    if args.month:
        target_sheets = [(args.month, wb.active if args.month is None else None)]
        # 해당 월 시트 찾기
        for name in wb.sheetnames:
            m = detect_month_from_sheet(name)
            if m == args.month:
                target_sheets = [(m, wb[name])]
                break
        else:
            target_sheets = [(args.month, wb.active)]
    else:
        # 전체 시트 - 월 이름 순으로 정렬
        target_sheets = []
        for name in wb.sheetnames:
            m = detect_month_from_sheet(name)
            if m:
                target_sheets.append((m, wb[name], name))
        target_sheets.sort(key=lambda x: x[0])

    # 각 시트 처리
    all_items = []
    all_unmatched = {}

    for entry in target_sheets:
        month_num, ws, sheet_name = entry
        print(f"\n📄 {sheet_name} ({args.year}년 {month_num}월) 처리 중...")
        items, unmatched = process_sheet(ws, args.year, month_num, cat_map, unmatched_cat)
        print(f"   → {len(items)}개 업무 발견")
        all_items.extend(items)
        for k, v in unmatched.items():
            all_unmatched[k] = all_unmatched.get(k, 0) + v

    # 미리보기
    print(f"\n{'='*60}")
    print(f"총 가져올 업무: {len(all_items)}개")
    print(f"{'='*60}")
    for item in all_items[:30]:
        cat_label = f"[{item['category_name']}]"
        print(f"  {item['date']}  {cat_label:<15} {item['title'][:40]}")
    if len(all_items) > 30:
        print(f"  ... 외 {len(all_items) - 30}개")

    if all_unmatched:
        print(f"\n⚠️  업무구분 미매칭 (미분류로 저장):")
        for proj, cnt in sorted(all_unmatched.items(), key=lambda x: -x[1]):
            print(f"   - '{proj}' ({cnt}건)")

    if args.dry_run:
        print("\n[dry-run] 실제 저장하지 않았습니다.")
        db.close()
        return

    confirm = input(f"\n위 {len(all_items)}개 업무를 DB에 저장할까요? (y/n): ").strip().lower()
    if confirm != "y":
        print("취소했습니다.")
        db.close()
        return

    for item in all_items:
        cat = item["category_obj"]
        task = Task(
            task_date=item["date"],
            category_id=cat.id if cat else unmatched_cat.id,
            title=item["title"],
            completed=False,
        )
        db.add(task)

    db.commit()
    db.close()
    print(f"\n✅ {len(all_items)}개 업무를 성공적으로 가져왔습니다!")
    if all_unmatched:
        print("   '미분류' 항목은 앱에서 업무구분을 직접 변경해주세요.")


if __name__ == "__main__":
    main()
